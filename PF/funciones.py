import openpyxl
from openpyxl.styles import Font
from conexion_BD import Conectar_BD
import os

def exportar_inventario_excel():
    try:
        conexion = Conectar_BD()
        cursor = conexion.cursor()
        sql = "SELECT id_producto, id_usuario, nombre_producto, unidades, precio_costo, precio_venta FROM inventario"
        cursor.execute(sql)
        inventario = cursor.fetchall()
        if not inventario:
            print("\n\t ⚠ No hay productos en el inventario para exportar ⚠")
            return
        ruta_excel = os.path.join(os.getcwd(), "inventario.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"
        columnas = ["ID Producto", "ID Usuario", "Nombre", "Unidades", "Precio Costo", "Precio Venta"]
        for col_num, columna in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=columna)
            cell.font = Font(bold=True)

        for fila_num, fila in enumerate(inventario, 2):
            for col_num, valor in enumerate(fila, 1):
                ws.cell(row=fila_num, column=col_num, value=valor)
        wb.save(ruta_excel)
        print(f"\n\t ✅ Inventario exportado correctamente ✅")
    except:
        print("\n\t ⚠ Error al exportar el inventario a Excel ⚠")


def borrarpantalla():
    import os
    os.system("cls")

def esperartecla():
    input("\n\t \U0001F552 Presiona ENTER para continuar...")
    
def menu_usurios():
   print("\n \t.:: Menu de usuarios::.. \n\t\t1.-Login  \n\t\t2.-Registro  \n\t\t3.-Salir ")
   opcion=input("\t Elige una opción: ").upper().strip() 
   return opcion

def menu_papeleria():
   print("\n \t .::  Menu Inventario::. \n\t1.- Agregar producto \n\t2.- Mostrar inventario \n\t3.- Modificar producto \n\t4.- Buscar producto \n\t5.- Eliminar producto \n\t6.- Productos bajos en stock \n\t7.- Regresar al menu de Inicio")
   opcion = input("\t\t Elige una opción (utiliza numero): ").strip()
   return opcion  

def menu_inicio():
    print("\n \t .::  Menu de Inicio::. \n\t1.- Inventario \n\t2.- ventas \n\t3.- Regresar al menu de usuarios")
    opcion = input("\t\t Elige una opción: ").upper().strip()
    return opcion  

def menu_ventas():
    print("\n \t .::  Menu Ventas::. \n\t1.- Realizar venta \n\t2.- Mostrar ventas \n\t3.- Modificar venta \n\t4.- Eliminar venta \n\t5.- Regresar al menu de Inicio")
    opcion = input("\t\t Elige una opción (utiliza numero): ").strip()
    return opcion
      